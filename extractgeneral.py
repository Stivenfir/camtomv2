import requests
import os
import json
import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.utils.exceptions import SheetTitleException

# Configura tu conexión
server = "172.16.10.77\\DBABC21"
database = "Repecev2005"
username = "Repecev2005"
password = ""

# Conexión
conn_str = (
    f"DRIVER={{ODBC Driver 17 for SQL Server}};"
    f"SERVER={server};"
    f"DATABASE={database};"
    f"UID={username};"
    f"PWD={password}"
)


def buscar_tipodoc(tipodoc):
    try:

        if tipodoc == "210":
            tipodoc = "47"

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Consulta con condición
        query = """
            SELECT *
            FROM IA_campostipodoc
            WHERE id_AItipodoc = ?
        """

        cursor.execute(query, (tipodoc,))
        rows = cursor.fetchall()

        return rows

    except Exception as e:
        print("Error en la conexión o consulta:", e)

    finally:
        if "conn" in locals():
            conn.close()


def ocr_docimportacion(tipodoc):

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    file_path = "C:\\Users\\aochoa\\Downloads\\SolicitudRegistro_generada_REG-50046191-20220328N.pdf"

    if tipodoc:
        # JSON como string
        json_data = {
            "document_type": "documento importacion",
            "document_description": "",
            "json_response": {
                "Indicador": f"{campostipodoc[0][3]}",
                "Cliente": f"{campostipodoc[1][3]}",
                "ClaseSolicitud": f"{campostipodoc[2][3]}",
                "RegimenImpo": f"{campostipodoc[3][3]}",
                "Exportador": f"{campostipodoc[4][3]}",
                "Ciudad": f"{campostipodoc[5][3]}",
                "AdminAduanas": f"{campostipodoc[6][3]}",
                "RImpoPaisOrigen": f"{campostipodoc[7][3]}",
                "RImpoPaisCompra": f"{campostipodoc[8][3]}",
                "ModoTransporte": f"{campostipodoc[9][3]}",
                "RImpoNoAprobacionDefinitiva": f"{campostipodoc[10][3]}",
                "RImpoNoAprobacion": f"{campostipodoc[11][3]}",
                "RImpoPuertoEmb": f"{campostipodoc[12][3]}",
                "RImpoFechaVencimiento": f"{campostipodoc[13][3]}",
                "RImpoFechaRadicacion": f"{campostipodoc[14][3]}",
                "RImpoFechaAprobacion": f"{campostipodoc[15][3]}",
                "RIMPOID": f"{campostipodoc[16][3]}",
                "PartidaArancelaria": f"{campostipodoc[17][3]}",
                "Referencia": f"{campostipodoc[18][3]}",
                "UnidadComercial": f"{campostipodoc[19][3]}",
                "cantidad": f"{campostipodoc[20][3]}",
                "PrecioUnitario": f"{campostipodoc[21][3]}",
                "paisorigen": f"{campostipodoc[22][3]}",
                "RefProductoMarca": f"{campostipodoc[23][3]}",
                "RefProductoDescripcion": f"{campostipodoc[24][3]}",
                "Descripcion_Mercancia": f"{campostipodoc[25][3]}",
                "PRODUCTO": f"{campostipodoc[26][3]}",
                "MARCA": f"{campostipodoc[27][3]}",
                "MODELO": f"{campostipodoc[28][3]}",
                "USO_O_DESTINO": f"{campostipodoc[29][3]}",
                "TIPO_DE_CORRIENTE": f"{campostipodoc[30][3]}",
                "NUMERO_DE_FASES": f"{campostipodoc[31][3]}",
                "POTENCIA": f"{campostipodoc[32][3]}",
                "MULTIPLICADOR_DE_VELOCIDAD": f"{campostipodoc[33][3]}",
            },
        }
    else:
        print("no está parametrizado ese tipo de documento.")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        "file_path": open(file_path, "rb"),
        "json_data": (None, json.dumps(json_data)),
    }

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        data = response.json()
        # Tomar el nombre del archivo con extensión
        hoja_nombre = os.path.basename(file_path)
        guardar_en_excel(data, hoja_nombre)
    else:
        print("❌ Error:", response.status_code)
        print(response.text)


def ocr_declimportacion(tipodoc):

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    file_path = "560373_4_Ingreso_Visado.pdf"

    if tipodoc:
        # JSON como string
        json_data = {
            "document_type": "documento importacion",
            "document_description": "",
            "json_response": {
                "ENCABEZADO_DO": f"{campostipodoc[0][3]}",
                "NombreImportador": f"{campostipodoc[1][3]}",
                "Cod.admon": f"{campostipodoc[2][3]}",
                "Cod.Dpto": f"{campostipodoc[3][3]}",
                "Cod.ciudad": f"{campostipodoc[4][3]}",
                "Clase_Importador": f"{campostipodoc[5][3]}",
                "ActEconomic": f"{campostipodoc[6][3]}",
                "TasaCambio": f"{campostipodoc[7][3]}",
                "Bultos": f"{campostipodoc[8][3]}",
                "Cod.Ingreso": f"{campostipodoc[9][3]}",
                "ManifCarga_FMM": f"{campostipodoc[10][3]}",
                "FechaManifCarga_FMM": f"{campostipodoc[11][3]}",
                "NoDocTransp": f"{campostipodoc[12][3]}",
                "FechaDocTransp": f"{campostipodoc[13][3]}",
                "Cod.Deposito": f"{campostipodoc[14][3]}",
                "Cod.Embalaje": f"{campostipodoc[15][3]}",
                "ModoTransporte": f"{campostipodoc[16][3]}",
                "Bandera": f"{campostipodoc[17][3]}",
                "NombreDeclarante": f"{campostipodoc[18][3]}",
                "NombreTransportador": f"{campostipodoc[19][3]}",
                "CodigoTransportador": f"{campostipodoc[20][3]}",
                "DeclImpo": [
                    {
                        "DeclImpoID": f"{campostipodoc[21][3]}",
                        "U.Com.": f"{campostipodoc[22][3]}",
                        "Factura": f"{campostipodoc[23][3]}",
                        "Fact.Fecha": f"{campostipodoc[24][3]}",
                        "P.Proce.": f"{campostipodoc[25][3]}",
                        "P.Origen": f"{campostipodoc[26][3]}",
                        "P.Compra": f"{campostipodoc[27][3]}",
                        "P.Arancelaria": f"{campostipodoc[28][3]}",
                        "Arancel.variable": f"{campostipodoc[29][3]}",
                        "Iva.Dif": f"{campostipodoc[30][3]}",
                        "Modalidad": f"{campostipodoc[31][3]}",
                        "Cod.Dpto.Dest.": f"{campostipodoc[32][3]}",
                        "Acuerdo": f"{campostipodoc[33][3]}",
                        "FormaPago": f"{campostipodoc[34][3]}",
                        "TipoImpo": f"{campostipodoc[35][3]}",
                        "TipoDecl.": f"{campostipodoc[36][3]}",
                        "NoCuotas": f"{campostipodoc[37][3]}",
                        "ValCuota": f"{campostipodoc[38][3]}",
                        "Periodicidad": f"{campostipodoc[39][3]}",
                        "Licencia": f"{campostipodoc[40][3]}",
                        "No": f"{campostipodoc[41][3]}",
                        "Año": f"{campostipodoc[42][3]}",
                        "NumeroPlan": f"{campostipodoc[43][3]}",
                        "CodigoProducto": f"{campostipodoc[44][3]}",
                        "NoSubpartidas": f"{campostipodoc[45][3]}",
                        "Casilla_35_DAV": f"{campostipodoc[46][3]}",
                        "Casilla_36_DAV": f"{campostipodoc[47][3]}",
                        "NombreExportador": f"{campostipodoc[48][3]}",
                        "Ciudad": f"{campostipodoc[49][3]}",
                        "Pais": f"{campostipodoc[50][3]}",
                        "Direccion": f"{campostipodoc[51][3]}",
                        "PesoBruto": f"{campostipodoc[52][3]}",
                        "PesoNeto": f"{campostipodoc[53][3]}",
                        "Cantidad": f"{campostipodoc[54][3]}",
                        "FobUS": f"{campostipodoc[55][3]}",
                        "FletesUS": f"{campostipodoc[56][3]}",
                        "SegurosUS": f"{campostipodoc[57][3]}",
                        "O.gastos": f"{campostipodoc[58][3]}",
                        "Sumatoria": f"{campostipodoc[59][3]}",
                        "AduanaUS": f"{campostipodoc[60][3]}",
                        "%Arancel": f"{campostipodoc[61][3]}",
                        "T.Liq.Arancel": f"{campostipodoc[62][3]}",
                        "%Iva": f"{campostipodoc[63][3]}",
                        "T.LiqIVA": f"{campostipodoc[64][3]}",
                        "Ajuste": f"{campostipodoc[65][3]}",
                        "Pago_total": f"{campostipodoc[66][3]}",
                        "%.Salvaguardia": f"{campostipodoc[67][3]}",
                        "T.Liq.Salvaguardia": f"{campostipodoc[68][3]}",
                        "%.Compensatorios": f"{campostipodoc[69][3]}",
                        "T.Liq.Compensatorio": f"{campostipodoc[70][3]}",
                        "%.Antidumping": f"{campostipodoc[71][3]}",
                        "T.Liq.Antidumpin": f"{campostipodoc[72][3]}",
                        "%.Sancio": f"{campostipodoc[73][3]}",
                        "T.Liq.Sancion": f"{campostipodoc[74][3]}",
                        "%.Rescate": f"{campostipodoc[75][3]}",
                        "T.Liq.Rescate": f"{campostipodoc[76][3]}",
                        "CANTIDAD_SERIALES_CARGADOS_A_LA_DECLARACION": f"{campostipodoc[77][3]}",
                        "ANTIDUMPING:": f"{campostipodoc[78][3]}",
                        "Ley_2277": f"{campostipodoc[79][3]}",
                        "SIC": f"{campostipodoc[80][3]}",
                        "ANTICIPADA": f"{campostipodoc[81][3]}",
                        "DESCRIPCION": f"{campostipodoc[82][3]}",
                    }
                ],
                "Totales": {
                    "PesoBruto": f"{campostipodoc[83][3]}",
                    "PesoNeto": f"{campostipodoc[84][3]}",
                    "Cantidad": f"{campostipodoc[85][3]}",
                    "FobUS": f"{campostipodoc[86][3]}",
                    "FletesUS": f"{campostipodoc[87][3]}",
                    "SegurosUS": f"{campostipodoc[88][3]}",
                    "O.gastos": f"{campostipodoc[89][3]}",
                    "Sumatoria": f"{campostipodoc[90][3]}",
                    "AduanaUS": f"{campostipodoc[91][3]}",
                    "T.Liq.Arancel": f"{campostipodoc[92][3]}",
                    "T.LiqIVA": f"{campostipodoc[93][3]}",
                    "T.Ajuste": f"{campostipodoc[94][3]}",
                    "Pago_Total": f"{campostipodoc[95][3]}",
                    "T.Liq.Salvaguardia": f"{campostipodoc[96][3]}",
                    "T.Liq.Compensatorios": f"{campostipodoc[97][3]}",
                    "T.Liq.Antidumpin": f"{campostipodoc[98][3]}",
                    "T.Liq.Sancion": f"{campostipodoc[99][3]}",
                    "T.Liq.Rescate": "",
                },
            },
        }
    else:
        print("no está parametrizado ese tipo de documento.")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        "file_path": open(file_path, "rb"),
        "json_data": (None, json.dumps(json_data)),
    }

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        data = response.json()
        # Tomar el nombre del archivo con extensión
        hoja_nombre = os.path.basename(file_path)
        guardar_en_excel(data, hoja_nombre)
    else:
        print("❌ Error:", response.status_code)
        print(response.text)


def normalizar_fecha(fecha_raw):
    import re

    if not fecha_raw:
        return ""
    fecha_raw = fecha_raw.strip()
    # Ya está en formato deseado
    if re.match(r"\d{2}/\d{2}/\d{4}", fecha_raw):
        return fecha_raw

    partes = fecha_raw.replace(".", "").split()
    if len(partes) != 3:
        return fecha_raw  # No se puede procesar

    dia, mes_abrev, anio = partes
    meses = {
        "JAN": "01",
        "FEB": "02",
        "MAR": "03",
        "APR": "04",
        "MAY": "05",
        "JUN": "06",
        "JUL": "07",
        "AUG": "08",
        "SEP": "09",
        "OCT": "10",
        "NOV": "11",
        "DEC": "12",
    }

    mes = meses.get(mes_abrev.upper())
    if not mes:
        return fecha_raw  # mes inválido

    return f"{dia.zfill(2)}/{mes}/{anio}"


def t(tipo, nullable=False, **kwargs):
    # La API espera tipos en MAYÚSCULAS: STRING, OBJECT, ARRAY, INTEGER...
    d = {"type": tipo.upper()}
    if nullable:
        d["nullable"] = True  # <- así, NO como ["string","null"]
    d.update(kwargs)
    return d


def ocr_factura(file_path, tipodoc):

    if tipodoc == "210":
        tipodoc = "47"

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL&json_schema=True"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    if campostipodoc:
        # JSON como string
        json_data = {
            "document_type": "Factura comercial",
            "document_description": """
El documento es una Factura Comercial Internacional de comercio exterior. Contiene datos del exportador e importador, Incoterms, logística y una tabla detallada de mercancías que puede incluir clasificaciones aduaneras (Partida Arancelaria / HS Code). Imporante tomar en cuenta que los valores monetarios deben interpretarse asumiendo la coma como separador decimal y así escribirse en el schema,por ejemplo 1.543,00 -> 1543.00, sin usar puntos, ni ningun otro separador para milesimas
 
**Reglas Críticas de Extracción de Ítems:**
1. Enumeración: Respeta estrictamente la literalidad del documento. Si la tabla de productos no tiene una columna impresa con números de ítem (ej: 1, 2, 3...), no generes una secuencia artificial; deja el campo de enumeración vacío.
2. Posición de Orden de Compra (order_position): Este campo es condicional. Solo debes extraer valor si existe una columna explícita en el detalle de productos etiquetada como 'Orden de Compra', 'PO' o similar vinculada a cada línea. Si la Orden de Compra solo aparece en la cabecera general del documento y no por producto, el campo order_position de los ítems debe quedar nulo/vacío.
 
 
**IMPORTANTE - Orden de Compra:**
1. ENCABEZADO: Buscar el número principal de orden de compra en el encabezado del documento (etiquetas: 'Orden', 'Purchase Order', 'PO', 'Pedido'). Este va en purchase_order.number_po.
2. TABLA DE PRODUCTOS: Si existe una columna 'Orden Compra', 'OC', 'PO' en la tabla de productos, los valores de esa columna (ej: 'ODC 106725', 'ODC 105850') van en el campo 'order_position' de cada item. NO confundir con 'reference' que es para SKU/código de producto.
""",
            "json_response": {
               "$schema": "http://json-schema.org/draft-07/schema#",
    "title": "Factura",
    "type": "object",
    "properties": {
        "factura": {
            "type": "object",
            "properties": {
                "invoiceNumber": {
                    "type": "string",
                    "description": "numero/id de la factura",
                    "pattern": "^[A-Z0-9_-]+$",
                    "minLength": 1,
                    "maxLength": 50,
                    "example": "INV-2025-0001",
                },
                "invoiceDate": {
                    "type": "string",
                    "description": "Fecha de emisión de la factura. Muchas veces se encuentra de manera diferente en los documentos. Puede ser: '25. JUL 2025', 'JUL 25, 2025', '25 JUL 2025', etc. El formato final debe ser DD/MM/YYYY",
                    "pattern": "^[0-9]{2}/[0-9]{2}/[0-9]{4}$",
                    "minLength": 10,
                    "maxLength": 11,
                    "example": "01/01/2025",
                },
                "incoterm": {
                    "type": "string",
                    "description": "Condición de entrega según Incoterms (por ejemplo, FOB, CIF, DDP)",
                    "pattern": "^[A-Z]{3}$",
                    "minLength": 3,
                    "maxLength": 3,
                    "example": "FOB",
                },
                "currency": {
                    "type": "string",
                    "description": "Código de la moneda de negociación en la que está expresada la factura. Debe ser un código ISO 4217 de 3 letras.",
                    "pattern": "^[A-Z]{3}$",
                    "minLength": 3,
                    "maxLength": 3,
                    "example": "USD",
                },
                "amount": {
                    "type": "string",
                    "description": "Subtotal de la factura antes de impuestos y cargos adicionales. ",
                    "pattern": "^\\d+(?:,\\d{1,2})?$",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "1000,00",
                },
                "total": {
                    "type": "string",
                    "description": "Total de la factura después de impuestos, flete y cargos adicionales.",
                    "pattern": "^\\d+(?:,\\d{1,2})?$",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "1000,00",
                },
                "freight_cost": {
                    "type": "string",
                    "description": "Costo del flete asociado a la factura, si aplica.",
                    "pattern": "^\\d+(?:,\\d{1,2})?$",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "1000,00",
                },
                "insurance": {
                    "type": "string",
                    "description": "Costo del seguro asociado a la factura, si aplica.",
                    "pattern": "^\\d+(?:,\\d{1,2})?$",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "1000,00",
                },
            },
            "required": [
                "invoiceNumber",
                "invoiceDate",
                "incoterm",
                "currency",
                "amount",
                "total",
            ],
        },
        "purchase_order": {
            "type": "object",
            "description": "Información de la orden de compra asociada a la factura. Buscar en el encabezado o sección de referencias del documento. Etiquetas comunes: 'Purchase Order', 'PO', 'P.O.', 'Orden de Compra', 'O.C.', 'OC', 'No. de Pedido', 'Pedido', 'Customer PO', 'Your Order', 'Order No.', 'Ref. Cliente'.",
            "properties": {
                "number_po": {
                    "type": ["string", "null"],
                    "description": "Número o código de la orden de compra del cliente. Buscar cerca de etiquetas como: 'PO Number', 'P.O. No.', 'Purchase Order', 'Orden de Compra', 'O.C.', 'Your Order No.', 'Customer PO', 'Pedido No.', 'Order Reference'. Usualmente aparece en el encabezado de la factura junto a la información de referencia. Puede tener formatos como: PO123456, 4500012345, OC-2024-001. Retornar null SOLO si definitivamente no existe ninguna referencia a orden de compra en el documento.",
                    "pattern": "^[A-Z0-9_-]+$",
                    "minLength": 1,
                    "maxLength": 50,
                    "example": "PO123456",
                },
                "date_po": {
                    "type": "string",
                    "nullable": True,
                    "description": "Fecha de la orden de compra. Buscar cerca de la etiqueta de orden de compra, puede aparecer como 'PO Date', 'Order Date', 'Fecha O.C.', 'Fecha de Pedido'. El formato final debe ser DD/MM/YYYY. Retornar null si no hay fecha específica para la orden de compra.",
                    "pattern": "^[0-9]{2}/[0-9]{2}/[0-9]{4}$",
                    "minLength": 10,
                    "maxLength": 11,
                    "example": "01/01/2025",
                },
                "position_po": {
                    "type": "string",
                    "nullable": True,
                    "description": "Posición o línea dentro de la orden de compra. Buscar en columnas como 'PO Line', 'Item', 'Pos.', 'Línea OC'. Retornar null si no existe esta información.",
                    "pattern": "^[A-Z0-9_-]+$",
                    "minLength": 1,
                    "maxLength": 10,
                    "example": "001",
                },
            },
        },
        "vendor": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "nullable": True,
                    "description": "Nombre del vendedor. E.g. Global Export Ltd.. Deberá ser null si no se encuentra explicitamente en el documento.",
                    
                    "minLength": 1,
                    "maxLength": 50,
                    "example": "John Doe Inc.",
                },
                "address": {
                    "type": "string",
                    "nullable": True,
                    "description": "Dirección del vendedor. E.g. 123 Export Rd, Hamburg, Germany. Deberá ser null si no se encuentra explicitamente en el documento.",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "123 Main St, Anytown, USA",
                },
                "legal_name": {
                    "type": "string",
                    "nullable": True,
                    "description": "Razón social del exportador o vendedor. E.g. Global Export Limited. Deberá ser null si no se encuentra explicitamente en el documento.",
                    
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "Global Export Limited",
                },
            },
        },
        "customer": {
            "type": "object",
            "properties": {
                "name": {
                    "type": "string",
                    "nullable": True,
                    "description": "Nombre del cliente. E.g. John Doe Inc.. Deberá ser null si no se encuentra explicitamente en el documento.",
                    
                    "minLength": 1,
                    "maxLength": 50,
                    "example": "John Doe Inc.",
                },
                "address": {
                    "type": "string",
                    "nullable": True,
                    "description": "Dirección del cliente. E.g. 123 Export Rd, Hamburg, Germany. Deberá ser null si no se encuentra explicitamente en el documento.",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "123 Main St, Anytown, USA",
                },
            },
        },
        "discharge": {
            "type": "object",
            "properties": {
                "address": {
                    "type": "string",
                    "nullable": True,
                    "description": "Dirección final del destinatario o entrega. E.g. 789 Delivery Blvd, Monterrey, Mexico.. Deberá ser null si no se encuentra explicitamente en el documento.",
                    
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "789 Delivery Blvd, Monterrey, Mexico",
                },
                "type": {
                    "type": "string",
                    "nullable": True,
                    "description": "Modo de transporte. Deberá ser null si no se encuentra explicitamente en el documento.",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "Aéreo",
                },
                "date": {
                    "type": "string",
                    "nullable": True,
                    "description": "Fecha de entrega o descarga estimada. en formato DD/MM/YYYY. E.g. 10/03/2024. Deberá ser null si no se encuentra explicitamente en el documento.",
                    "pattern": "^[0-9]{2}/[0-9]{2}/[0-9]{4}$",
                    "minLength": 1,
                    "maxLength": 100,
                    "example": "10/03/2024",
                },
            },
        },
        "delivery_place": {
            "type": "string",
            "nullable": True,
            "description": "Lugar de entrega. Puede ser ciudad o puerto. E.g. Veracruz, Hamburg. Deberá ser null si no se encuentra explicitamente en el documento.",
            "minLength": 1,
            "maxLength": 100,
            "example": "Veracruz",
        },
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "item_position": {
                        "type": "string",
                        "description": "Posición del ítem en la factura. E.g. 001. Deberá ser null si no se encuentra explicitamente en el documento.",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "001",
                    },
                    "order_position": {
                        "type": "string",
                        "description": "Número o código de la orden de compra asociada a este ítem/producto específico. IMPORTANTE: Buscar en columnas con encabezados como 'Orden Compra', 'OC', 'O.C.', 'PO', 'Purchase Order', 'Pedido', 'No. Pedido'. Los valores típicos tienen formato como 'ODC 106725', 'OC-12345', 'PO123456'. Este campo captura el código de orden de compra POR LÍNEA de producto. NO es lo mismo que 'reference'. Retornar null solo si no existe columna de orden de compra en la tabla.",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "ODC 106725",
                    },
                    "reference": {
                        "type": "string",
                        "description": "Código SKU, número de parte o referencia interna del producto (NO es la orden de compra). Buscar en columnas como 'SKU', 'Part No.', 'Código', 'Ref. Producto', 'Item Code'. Retornar null si no existe este tipo de referencia de producto.",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "PRD-12345",
                    },
                    "origin_country": {
                        "type": "string",
                        "description": "Priorizar el país de origen especificado a nivel ítem. Si no se encuentra explícitamente a nivel ítem, utilizar el país de origen indicado en la factura. Si no se menciona en ninguno de los dos casos, deberá ser null",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "México",
                    },
                    "brand": {
                        "type": "string",
                        "description": "Marca del item.\n Reglas estrictas:\n1) Si existe una columna específica para marca (i.e. Marca, Mark, Brand o algún término similar), usarla.\n2) Si no, analizar la descripción. Extraerla SOLO si es el viene la marca de forma explícita (precedida de 'Brand:', 'Mark', o algún término similar). REGLA DE EXCLUSIÓN: Si la descripción comienza con el nombre de una línea de producto, familia, modelo, tecnología o un nombre comercial compuesto (especialmente aquellos con guiones o números), NO extraerlo y dejar el campo vacío. Ejemplos genéricos de lo que NO se debe extraer: 'Galaxy' (Línea), 'Mustang' (Modelo), 'PlayStation' (Sub-marca), 'Wi-Fi' (Tecnología), 'Power-Shot' (Nombre compuesto).\n3) ANTI-ALUCINACIÓN: Si la marca no está escrita explícitamente en la línea del ítem, dejar vacío. No rellenar basándose en logotipos del encabezado o conocimiento externo.",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "Dell",
                    },
                    "description": {
                        "type": "string",
                        "description": "Descripción detallada del producto o servicio facturado. Deberá ser null si no se encuentra explicitamente en el documento.",
                        "minLength": 0,
                        "maxLength": 1000,
                        "example": "Laptop Dell XPS 13, Intel i7, 16GB RAM, 512GB SSD",
                    },
                    "year_manufacture": {
                        "type": "integer",
                        "description": "Año de fabricación del producto. Deberá ser null si no se encuentra explicitamente en el documento.",
                        "example": 2023,
                    },
                    "quantity": {
                        "type": "integer",
                        "description": "Cantidad de unidades del producto o servicio facturado.",
                        "minimum": 0,
                        "example": 3,
                    },
                    "unit": {
                        "type": "string",
                        "description": "Unidad de medida del producto o servicio (por ejemplo, pcs, kg, liters). Deberá ser null si no se encuentra explicitamente en el documento.",
                        "minLength": 0,
                        "maxLength": 20,
                        "example": "liters",
                    },
                    "unitPrice": {
                        "type": "string",
                        "description": "Precio unitario del producto o servicio (por ejemplo, 19954,99). Deberá ser null si no se encuentra explicitamente en el documento.",
                        "minLength": 1,
                        "maxLength": 20,
                        "pattern": "^\\d+(?:,\\d{1,2})?$",
                        "example": "19954,99",
                    },
                    "subTotal": {
                        "type": "string",
                        "description": "Subtotal del producto o servicio (por ejemplo, 599,97). Deberá ser null si no se encuentra explicitamente en el documento.",
                        "minLength": 1,
                        "maxLength": 20,
                        "pattern": "^\\d+(?:,\\d+)?$",
                        "example": "599,97",
                    },
                    "totalweight_kg": {
                        "type": ["string", "NULL"],
                        "description": "Peso bruto total del producto en kilogramos (por ejemplo, 1,5). Deberá ser null si no se encuentra explicitamente en el documento.",
                        "pattern": "^\\d+(?:,\\d{1,2})?$",
                        "example": "1,5",
                    },
                    "totalnetweight_kg": {
                        "type": "string",
                        "nullable": True,
                        "description": "Peso neto total del producto en kilogramos (por ejemplo, 1,2). Deberá ser null si no se encuentra explicitamente en el documento.",
                        "pattern": "^\\d+(?:,\\d{1,2})?$",
                        "example": "1,2",
                    },
                    "amount": {
                        "type": "string",
                        "description": "Valor total del producto o servicio (por ejemplo, 6599,97). Deberá ser null si no se encuentra explicitamente en el documento. Utilizar comas como separador decimal",
                        "pattern": "^\\d+(?:,\\d{1,2})?$",
                        "minLength": 1,
                        "maxLength": 100,
                        "example": "1000,00",
                    },
                },
                            "required": [
                                "description",
                                "quantity",
                                "unitPrice",
                                "subTotal",
                                "amount",
                            ],
                        },
                        "minItems": 1,
                    },
                },
                "required": ["factura", "items"]
            }
            
        }
    else:
        print(f"no está parametrizado ese tipo de documento. {tipodoc}")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        "file_path": open(file_path, "rb"),
        "json_data": (None, json.dumps(json_data)),
    }

 

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        hoja_nombre = os.path.basename(file_path)

        response_json = response.json()
        print(response_json)

        # OJO: según la API, a veces viene en "document_data"
        doc = response_json.get("document_data", {})

        # ✅ Normaliza fechas específicas en la respuesta (DENTRO del if)
        # Factura
        if isinstance(doc, dict) and "factura" in doc and isinstance(doc["factura"], dict):
            doc["factura"]["invoiceDate"] = normalizar_fecha(doc["factura"].get("invoiceDate", ""))

        # Purchase order
        if isinstance(doc, dict) and "purchase_order" in doc and isinstance(doc["purchase_order"], dict):
            doc["purchase_order"]["date_po"] = normalizar_fecha(doc["purchase_order"].get("date_po", ""))

        # Discharge
        if isinstance(doc, dict) and "discharge" in doc and isinstance(doc["discharge"], dict):
            doc["discharge"]["date"] = normalizar_fecha(doc["discharge"].get("date", ""))

        # Si quieres guardar en excel lo normalizado, guarda doc o response_json
        # response_json["document_data"] = doc
        # guardar_en_excel(response_json, hoja_nombre)

        return response

    else:
        print("❌ Error:", response.status_code)
        print(response.text)
        return response


"""
"invoice": {
                    "id": f"{campostipodoc[0][3]}",
                    "date": f"{campostipodoc[1][3]}",
                    "incoterm": f"{campostipodoc[2][3]}",
                    "currency": f"{campostipodoc[3][3]}",
                    "amount": f"{campostipodoc[4][3]}",
                    "total": f"{campostipodoc[5][3]}",
                    "freight_cost": f"{campostipodoc[6][3]}",
                    "insurance": f"{campostipodoc[7][3]}"
                },
                "purchase_order": {
                    "number": f"{campostipodoc[8][3]}",
                    "date": f"{campostipodoc[9][3]}",
                    "position": f"{campostipodoc[10][3]}"
                },
                "vendor": {
                    "name": f"{campostipodoc[11][3]}",
                    "address": f"{campostipodoc[12][3]}",
                    "legal_name": f"{campostipodoc[13][3]}"
                },
                "customer": {
                    "name": f"{campostipodoc[14][3]}",
                    "address": f"{campostipodoc[15][3]}"
                },
                "discharge": {
                    "address": f"{campostipodoc[16][3]}",
                    "type": f"{campostipodoc[17][3]}",
                    "date": f"{campostipodoc[18][3]}"
                },
                "delivery_place": f"{campostipodoc[19][3]}",
                "items": [
                    {   
                        "purchaseorder_number_item": '',
                        "order_date": '',
                        "order_position": f"{campostipodoc[20][3]}",
                        "reference": f"{campostipodoc[21][3]}",
                        "origin_country": f"{campostipodoc[22][3]}",
                        "brand": f"{campostipodoc[23][3]}",
                        "description": f"{campostipodoc[24][3]}",
                        "year_manufacture": f"{campostipodoc[25][3]}",
                        "quantity": f"{campostipodoc[26][3]}",
                        "unit": f"{campostipodoc[27][3]}",
                        "unit_price": f"{campostipodoc[28][3]}",
                        "subtotal": f"{campostipodoc[29][3]}",
                        "total_amount": f"{campostipodoc[32][3]}",
                        "gross_weight_kg": f"{campostipodoc[30][3]}",
                        "net_weight_kg": f"{campostipodoc[31][3]}",
                    }
                ]

"""


def guardar_en_excel(data, hoja_nombre):
    excel_path = "C:\\CAMTOM\\resultado_facturas_0725.xlsx"

    # Convertir a dataframe plano (items se separan si están en lista)
    registros = []

    def descomponer_json(obj, base=""):
        if isinstance(obj, dict):
            for k, v in obj.items():
                descomponer_json(v, f"{base}.{k}" if base else k)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                descomponer_json(item, f"{base}[{i}]")
        else:
            registros.append((base, obj))

    descomponer_json(data)
    df = pd.DataFrame(registros, columns=["Campo", "Valor"])

    # Limpiar nombre hoja si es muy largo o contiene caracteres inválidos
    hoja_nombre = (
        hoja_nombre.replace("\\", "_").replace("/", "_").replace(" ", "_")[-31:]
    )

    # Crear nuevo Excel o abrir existente
    if not os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=hoja_nombre, index=False)
    else:
        book = load_workbook(excel_path)
        with pd.ExcelWriter(
            excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            # Eliminar hoja si ya existe
            if hoja_nombre in book.sheetnames:
                del book[hoja_nombre]
            try:
                df.to_excel(writer, sheet_name=hoja_nombre, index=False)
            except SheetTitleException:
                df.to_excel(writer, sheet_name="Factura", index=False)

    print(f"📄 Datos guardados en hoja '{hoja_nombre}' de {excel_path}")


# ocr_factura('C:\\CAMTOM\\facturaspruebas0412\\FACTURACOMERCIAL-DrummondFacturanumero.pdf', '210')
# ocr_docimportacion("1")
# ocr_declimportacion("2")
