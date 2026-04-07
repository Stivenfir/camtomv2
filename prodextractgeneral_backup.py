import requests
import os
import json
import pandas as pd
import pyodbc
from openpyxl import load_workbook
from openpyxl.utils.exceptions import SheetTitleException

# Configura tu conexión
server = "172.16.10.16\\DBABC21"
database = "Repecev2005_H"
username = "Repecev2005"
password = ""

# Conexión
conn_str = (
    f'DRIVER={{ODBC Driver 17 for SQL Server}};'
    f'SERVER={server};'
    f'DATABASE={database};'
    f'UID={username};'
    f'PWD={password}'
)

def buscar_tipodoc(tipodoc):
    try:

        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()

        # Consulta con condición
        query = """
            SELECT *
            FROM IA_campostipodoc
            WHERE id_AItipodoc = ?
        """

        cursor.execute(query, (tipodoc,))
        rows = cursor.fetchall()

        return rows

    except Exception as e:
        print("Error en la conexión o consulta:", e)

    finally:
        if 'conn' in locals():
            conn.close()

def ocr_docimportacion(tipodoc):

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    file_path = '14584369.pdf'

    if tipodoc:
        # JSON como string
        json_data = {
            "document_type": "documento importacion",
            "document_description": "",
            "json_response": {
                "Indicador": f"{campostipodoc[0][3]}",
                "Cliente": f"{campostipodoc[1][3]}",
                "ClaseSolicitud": f"{campostipodoc[2][3]}",
                "RegimenImpo": f"{campostipodoc[3][3]}",
                "Exportador": f"{campostipodoc[4][3]}",
                "Ciudad": f"{campostipodoc[5][3]}",
                "AdminAduanas": f"{campostipodoc[6][3]}",
                "RImpoPaisOrigen": f"{campostipodoc[7][3]}",
                "RImpoPaisCompra": f"{campostipodoc[8][3]}",
                "ModoTransporte": f"{campostipodoc[9][3]}",
                "RImpoNoAprobacionDefinitiva": f"{campostipodoc[10][3]}",
                "RImpoNoAprobacion": f"{campostipodoc[11][3]}",
                "RImpoPuertoEmb": f"{campostipodoc[12][3]}",
                "RImpoFechaVencimiento": f"{campostipodoc[13][3]}",
                "RImpoFechaRadicacion": f"{campostipodoc[14][3]}",
                "RImpoFechaAprobacion": f"{campostipodoc[15][3]}",
                "RIMPOID": f"{campostipodoc[16][3]}",
                "PartidaArancelaria": f"{campostipodoc[17][3]}",
                "Referencia": f"{campostipodoc[18][3]}",
                "UnidadComercial": f"{campostipodoc[19][3]}",
                "cantidad": f"{campostipodoc[20][3]}",
                "PrecioUnitario": f"{campostipodoc[21][3]}",
                "paisorigen": f"{campostipodoc[22][3]}",
                "RefProductoMarca": f"{campostipodoc[23][3]}",
                "RefProductoDescripcion": f"{campostipodoc[24][3]}",
                "Descripcion_Mercancia": f"{campostipodoc[25][3]}",
                "PRODUCTO": f"{campostipodoc[26][3]}",
                "MARCA": f"{campostipodoc[27][3]}",
                "MODELO": f"{campostipodoc[28][3]}",
                "USO_O_DESTINO": f"{campostipodoc[29][3]}",
                "TIPO_DE_CORRIENTE": f"{campostipodoc[30][3]}",
                "NUMERO_DE_FASES": f"{campostipodoc[31][3]}",
                "POTENCIA": f"{campostipodoc[32][3]}",
                "MULTIPLICADOR_DE_VELOCIDAD": f"{campostipodoc[33][3]}",
            }
        }
    else:
        print("no está parametrizado ese tipo de documento.")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        'file_path': open(file_path, 'rb'),
        'json_data': (None, json.dumps(json_data))
    }

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        data = response.json()
        # Tomar el nombre del archivo con extensión
        hoja_nombre = os.path.basename(file_path)
        guardar_en_excel(data, hoja_nombre)
    else:
        print("❌ Error:", response.status_code)
        print(response.text)

def ocr_declimportacion(tipodoc):

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    file_path = '560373_4_Ingreso_Visado.pdf'

    if tipodoc:
        # JSON como string
        json_data = {
            "document_type": "documento importacion",
            "document_description": "",
            "json_response": {
                "ENCABEZADO_DO": f"{campostipodoc[0][3]}",
                "NombreImportador": f"{campostipodoc[1][3]}",
                "Cod.admon": f"{campostipodoc[2][3]}",
                "Cod.Dpto": f"{campostipodoc[3][3]}",
                "Cod.ciudad": f"{campostipodoc[4][3]}",
                "Clase_Importador": f"{campostipodoc[5][3]}",
                "ActEconomic": f"{campostipodoc[6][3]}",
                "TasaCambio": f"{campostipodoc[7][3]}",
                "Bultos": f"{campostipodoc[8][3]}",
                "Cod.Ingreso": f"{campostipodoc[9][3]}",
                "ManifCarga_FMM": f"{campostipodoc[10][3]}",
                "FechaManifCarga_FMM": f"{campostipodoc[11][3]}",
                "NoDocTransp": f"{campostipodoc[12][3]}",
                "FechaDocTransp": f"{campostipodoc[13][3]}",
                "Cod.Deposito": f"{campostipodoc[14][3]}",
                "Cod.Embalaje": f"{campostipodoc[15][3]}",
                "ModoTransporte": f"{campostipodoc[16][3]}",
                "Bandera": f"{campostipodoc[17][3]}",
                "NombreDeclarante": f"{campostipodoc[18][3]}",
                "NombreTransportador": f"{campostipodoc[19][3]}",
                "CodigoTransportador": f"{campostipodoc[20][3]}",
                "DeclImpo": [
                    {
                        "DeclImpoID": f"{campostipodoc[21][3]}",
                        "U.Com.": f"{campostipodoc[22][3]}",
                        "Factura": f"{campostipodoc[23][3]}",
                        "Fact.Fecha": f"{campostipodoc[24][3]}",
                        "P.Proce.": f"{campostipodoc[25][3]}",
                        "P.Origen": f"{campostipodoc[26][3]}",
                        "P.Compra": f"{campostipodoc[27][3]}",
                        "P.Arancelaria": f"{campostipodoc[28][3]}",
                        "Arancel.variable": f"{campostipodoc[29][3]}",
                        "Iva.Dif": f"{campostipodoc[30][3]}",
                        "Modalidad": f"{campostipodoc[31][3]}",
                        "Cod.Dpto.Dest.": f"{campostipodoc[32][3]}",
                        "Acuerdo": f"{campostipodoc[33][3]}",
                        "FormaPago": f"{campostipodoc[34][3]}",
                        "TipoImpo": f"{campostipodoc[35][3]}",
                        "TipoDecl.": f"{campostipodoc[36][3]}",
                        "NoCuotas": f"{campostipodoc[37][3]}",
                        "ValCuota": f"{campostipodoc[38][3]}",
                        "Periodicidad": f"{campostipodoc[39][3]}",
                        "Licencia": f"{campostipodoc[40][3]}",
                        "No": f"{campostipodoc[41][3]}",
                        "Año": f"{campostipodoc[42][3]}",
                        "NumeroPlan": f"{campostipodoc[43][3]}",
                        "CodigoProducto": f"{campostipodoc[44][3]}",
                        "NoSubpartidas": f"{campostipodoc[45][3]}",
                        "Casilla_35_DAV": f"{campostipodoc[46][3]}",
                        "Casilla_36_DAV": f"{campostipodoc[47][3]}",
                        "NombreExportador": f"{campostipodoc[48][3]}",
                        "Ciudad": f"{campostipodoc[49][3]}",
                        "Pais": f"{campostipodoc[50][3]}",
                        "Direccion": f"{campostipodoc[51][3]}",
                        "PesoBruto": f"{campostipodoc[52][3]}",
                        "PesoNeto": f"{campostipodoc[53][3]}",
                        "Cantidad": f"{campostipodoc[54][3]}",
                        "FobUS": f"{campostipodoc[55][3]}",
                        "FletesUS": f"{campostipodoc[56][3]}",
                        "SegurosUS": f"{campostipodoc[57][3]}",
                        "O.gastos": f"{campostipodoc[58][3]}",
                        "Sumatoria": f"{campostipodoc[59][3]}",
                        "AduanaUS": f"{campostipodoc[60][3]}",
                        "%Arancel": f"{campostipodoc[61][3]}",
                        "T.Liq.Arancel": f"{campostipodoc[62][3]}",
                        "%Iva": f"{campostipodoc[63][3]}",
                        "T.LiqIVA": f"{campostipodoc[64][3]}",
                        "Ajuste": f"{campostipodoc[65][3]}",
                        "Pago_total": f"{campostipodoc[66][3]}",
                        "%.Salvaguardia": f"{campostipodoc[67][3]}",
                        "T.Liq.Salvaguardia": f"{campostipodoc[68][3]}",
                        "%.Compensatorios": f"{campostipodoc[69][3]}",
                        "T.Liq.Compensatorio": f"{campostipodoc[70][3]}",
                        "%.Antidumping": f"{campostipodoc[71][3]}",
                        "T.Liq.Antidumpin": f"{campostipodoc[72][3]}",
                        "%.Sancio": f"{campostipodoc[73][3]}",
                        "T.Liq.Sancion": f"{campostipodoc[74][3]}",
                        "%.Rescate": f"{campostipodoc[75][3]}",
                        "T.Liq.Rescate": f"{campostipodoc[76][3]}",
                        "CANTIDAD_SERIALES_CARGADOS_A_LA_DECLARACION": f"{campostipodoc[77][3]}",
                        "ANTIDUMPING:": f"{campostipodoc[78][3]}",
                        "Ley_2277": f"{campostipodoc[79][3]}",
                        "SIC": f"{campostipodoc[80][3]}",
                        "ANTICIPADA": f"{campostipodoc[81][3]}",
                        "DESCRIPCION": f"{campostipodoc[82][3]}",
                    }
                ],
                "Totales": {
                    "PesoBruto": f"{campostipodoc[83][3]}",
                    "PesoNeto": f"{campostipodoc[84][3]}",
                    "Cantidad": f"{campostipodoc[85][3]}",
                    "FobUS": f"{campostipodoc[86][3]}",
                    "FletesUS": f"{campostipodoc[87][3]}",
                    "SegurosUS": f"{campostipodoc[88][3]}",
                    "O.gastos": f"{campostipodoc[89][3]}",
                    "Sumatoria": f"{campostipodoc[90][3]}",
                    "AduanaUS": f"{campostipodoc[91][3]}",
                    "T.Liq.Arancel": f"{campostipodoc[92][3]}",
                    "T.LiqIVA": f"{campostipodoc[93][3]}",
                    "T.Ajuste": f"{campostipodoc[94][3]}",
                    "Pago_Total": f"{campostipodoc[95][3]}",
                    "T.Liq.Salvaguardia": f"{campostipodoc[96][3]}",
                    "T.Liq.Compensatorios": f"{campostipodoc[97][3]}",
                    "T.Liq.Antidumpin": f"{campostipodoc[98][3]}",
                    "T.Liq.Sancion": f"{campostipodoc[99][3]}",
                    "T.Liq.Rescate": ""
                }
            }
        }
    else:
        print("no está parametrizado ese tipo de documento.")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        'file_path': open(file_path, 'rb'),
        'json_data': (None, json.dumps(json_data))
    }

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        data = response.json()
        # Tomar el nombre del archivo con extensión
        hoja_nombre = os.path.basename(file_path)
        guardar_en_excel(data, hoja_nombre)
    else:
        print("❌ Error:", response.status_code)
        print(response.text)

def ocr_factura(file_path, tipodoc):

    campostipodoc = buscar_tipodoc(tipodoc)

    # URL de la API
    url = "https://api.camtomx.com/api/v3/camtomdocs/extract?country_code=COL"

    # Encabezados
    headers = {
        "accept": "application/json",
        "Authorization": "Bearer sk_6345e9288e81f89290ac68a279f6e22c1804fb74f7c5758f8b3a0235f6af61e2",  # Reemplaza con tu token real
    }

    if campostipodoc:
        # JSON como string
        json_data = {
            "document_type": "Factura comercial",
            "document_description": "Factura comercial internacional utilizada en operaciones de comercio exterior. Contiene datos de comprador, vendedor, condiciones de venta, productos, valores y logística.",
            "json_response": {
                "invoice": {
                    "id": f"{campostipodoc[0][3]}",
                    "date": f"{campostipodoc[1][3]}",
                    "incoterm": f"{campostipodoc[2][3]}",
                    "currency": f"{campostipodoc[3][3]}",
                    "amount": f"{campostipodoc[4][3]}",
                    "total": f"{campostipodoc[5][3]}",
                    "freight_cost": f"{campostipodoc[6][3]}",
                    "insurance": f"{campostipodoc[7][3]}"
                },
                "purchase_order": {
                    "number": f"{campostipodoc[8][3]}",
                    "date": f"{campostipodoc[9][3]}",
                    "position": f"{campostipodoc[10][3]}"
                },
                "vendor": {
                    "name": f"{campostipodoc[11][3]}",
                    "address": f"{campostipodoc[12][3]}",
                    "legal_name": f"{campostipodoc[13][3]}"
                },
                "customer": {
                    "name": f"{campostipodoc[14][3]}",
                    "address": f"{campostipodoc[15][3]}"
                },
                "discharge": {
                    "address": f"{campostipodoc[16][3]}",
                    "type": f"{campostipodoc[17][3]}",
                    "date": f"{campostipodoc[18][3]}"
                },
                "delivery_place": f"{campostipodoc[19][3]}",
                "items": [
                    {   
                        "purchaseorder_number_item": '',
                        "order_date": '',
                        "order_position": f"{campostipodoc[20][3]}",
                        "reference": f"{campostipodoc[21][3]}",
                        "origin_country": f"{campostipodoc[22][3]}",
                        "brand": f"{campostipodoc[23][3]}",
                        "description": f"{campostipodoc[24][3]}",
                        "year_manufacture": f"{campostipodoc[25][3]}",
                        "quantity": f"{campostipodoc[26][3]}",
                        "unit": f"{campostipodoc[27][3]}",
                        "unit_price": f"{campostipodoc[28][3]}",
                        "subtotal": f"{campostipodoc[29][3]}",
                        "total_amount": f"{campostipodoc[32][3]}",
                        "gross_weight_kg": f"{campostipodoc[30][3]}",
                        "net_weight_kg": f"{campostipodoc[31][3]}",
                    }
                ]
            }
        }
    else:
        print(f"no está parametrizado ese tipo de documento. {tipodoc}")

    print(json_data)

    # Enviar archivo y JSON a la API
    files = {
        'file_path': open(file_path, 'rb'),
        'json_data': (None, json.dumps(json_data))
    }

    response = requests.post(url, headers=headers, files=files)

    if response.status_code == 200:
        print("✅ Extracción exitosa")
        print(response)
        return response
    else:
        print("❌ Error:", response.status_code)
        print(response.text)
        return response

def guardar_en_excel(data, hoja_nombre):
    excel_path = "C:\\CAMTOM\\resultado_facturas.xlsx"

    # Convertir a dataframe plano (items se separan si están en lista)
    registros = []

    def descomponer_json(obj, base=""):
        if isinstance(obj, dict):
            for k, v in obj.items():
                descomponer_json(v, f"{base}.{k}" if base else k)
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                descomponer_json(item, f"{base}[{i}]")
        else:
            registros.append((base, obj))

    descomponer_json(data)
    df = pd.DataFrame(registros, columns=["Campo", "Valor"])

    # Limpiar nombre hoja si es muy largo o contiene caracteres inválidos
    hoja_nombre = hoja_nombre.replace("\\", "_").replace("/", "_").replace(" ", "_")[-31:]

    # Crear nuevo Excel o abrir existente
    if not os.path.exists(excel_path):
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=hoja_nombre, index=False)
    else:
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Eliminar hoja si ya existe
            if hoja_nombre in book.sheetnames:
                del book[hoja_nombre]
            try:
                df.to_excel(writer, sheet_name=hoja_nombre, index=False)
            except SheetTitleException:
                df.to_excel(writer, sheet_name="Factura", index=False)

    print(f"📄 Datos guardados en hoja '{hoja_nombre}' de {excel_path}")

#ocr_factura("1")
# ocr_docimportacion("1")
#ocr_declimportacion("2")